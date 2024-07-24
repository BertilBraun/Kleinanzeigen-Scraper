import json
import base64
import asyncio

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font

from pprint import pprint

from openai import AsyncOpenAI

from src.config import OPENAI_API_KEY
from src.scrape_functions import scrape_all_offer_links_from_search_url, scrape_offer_url
from src.defines import BASE_URL, CURRENT_OFFERS_FILE, DB_FILE, FILTERED_OUT_OFFERS_FILE, LLM_MODEL_ID
from src.types import DatabaseFactory, Entry, Offer
from src.util import dump_json, log_all_exceptions, timeblock


async def scrape_all_offers(all_offer_links: list[str]) -> list[Offer]:
    offer_futures = [scrape_offer_url(url) for url in all_offer_links]
    offers: list[Offer] = []
    for offer in asyncio.as_completed(offer_futures):
        with log_all_exceptions('while scraping offer'):
            offers.append(await offer)
    return offers


def load_database() -> list[Entry]:
    try:
        with open(DB_FILE, 'r') as file:
            return DatabaseFactory.from_json(json.load(file))
    except FileNotFoundError:
        return []


def partition_offers(
    all_current_offers: list[Offer], database_entries: list[Entry]
) -> tuple[list[Offer], list[tuple[Offer, Entry]], list[Entry]]:
    # partition into: new offers which are not yet in the database, offers which are already in the database but still in the current offers, and offers which are no longer in the current offers
    new_offers: list[Offer] = []

    for offer in all_current_offers:
        if not any(entry.metadata.offer.id == offer.id for entry in database_entries):
            new_offers.append(offer)

    old_offers: list[tuple[Offer, Entry]] = []
    sold_offers: list[Entry] = []

    for entry in database_entries:
        if not any(offer.id == entry.metadata.offer.id for offer in all_current_offers):
            sold_offers.append(entry)
        else:
            offer = next(offer for offer in all_current_offers if offer.id == entry.metadata.offer.id)
            old_offers.append((offer, entry))

    return new_offers, old_offers, sold_offers


def get_example_image() -> str:
    # load example_prompt_image.jpeg and convert to base64
    with open('example_prompt_image.jpeg', 'rb') as file:
        return base64.b64encode(file.read()).decode('utf-8')


async def extract_offer_details(offer: Offer) -> Entry | Offer:
    client = AsyncOpenAI(api_key=OPENAI_API_KEY)

    base64_example_image = get_example_image()

    response = await client.chat.completions.create(
        model=LLM_MODEL_ID,
        messages=[
            {
                'role': 'system',
                'content': """You are a helpful assistant that extracts information from eBay Kleinanzeigen related to Windsurf equipment and converts it into a specific JSON format. The types of equipment include sails, boards, masts, booms, full sets, full rigs, and accessories. 

If the information is not available or cannot be determined from the input, use "".

You should output the information in the following JSON format based on the type of equipment:

Sail:
```json
{
  "type": "sail",
  "size": "Size of the Sail in m²",
  "brand": "Name of the Brand and Model",
  "mast_length": "Length of the required Mast in cm",
  "boom_size": "Size of the required Boom in cm",
  "year": "Release Year",
  "state": "new, used, repaired, demaged, defective"
}
```

Board:
```json
{
  "type": "board",
  "size": "Dimensions of the Board",
  "brand": "Name of the Brand and Model",
  "board_type": "Freeride, Wave, Freestyle, Slalom, Formula, ...",
  "volume": "Volume in Liters",
  "year": "Release Year"
}
```

Mast:
```json
{
  "type": "mast",
  "brand": "Name of the Brand and Model",
  "length": "Length of the Mast in cm",
  "carbon": "Carbon Percentage",
  "rdm_or_sdm": "Either RDM or SDM"
}
```

Boom:
```json
{
  "type": "boom",
  "brand": "Name of the Brand and Model",
  "size": "Minimum and Maximum Size of the Boom in cm",
  "year": "Release Year"
}
```

Full Set:
```json
{
  "type": "full_set",
  "content_description": "Short description of what the set includes (e.g., Sail, Mast, Boom, Board, etc.)"
}
```

Full Rig:
```json
{
  "type": "full_rig",
  "sail": {
    "type": "sail",
    "size": "Size of the Sail in m²",
    "brand": "Name of the Brand and Model",
    "mast_length": "Length of the required Mast in cm",
    "boom_size": "Size of the required Boom in cm",
    "year": "Release Year",
    "state": "new, used, repaired, demaged, defective"
  },
  "mast": {
    "type": "mast",
    "brand": "Name of the Brand and Model",
    "length": "Length of the Mast in cm",
    "carbon": "Carbon Percentage",
    "rdm_or_sdm": "Either RDM or SDM"
  },
  "boom": {
    "type": "boom",
    "brand": "Name of the Brand and Model",
    "size": "Minimum and Maximum Size of the Boom in cm",
    "year": "Release Year"
  }
}
```

Accessory:
```json
{
  "type": "accessory",
  "accessory_type": "Mastfoot, Harness Lines, Fins, etc.",
}
```

If the type of equipment cannot be determined or is not relevant to usable windsurf equipment, use:
```json
{
  "type": "N/A"
}
```
This will be for items like child equipment, courses, toys, etc. which are not relevant to windsurfing.

""",
            },
            {
                'role': 'user',
                'content': [
                    {
                        'type': 'text',
                        'text': """As an example, let's extract the details of the following offer:
---

Convert the following offer into the appropriate JSON format:

Title: North Spectro 6.5 Surfsegel Windsurfen
Description: Segel mit wenigen Gebrauchsspuren. 2 Band-Camber als Profilgeber. Ein kleiner getapteter Cut im Unterliek. gerne auch mit Carbonmast + 20€""",
                    },
                    {
                        'type': 'image_url',
                        'image_url': {
                            'url': f'data:image/png;base64,{base64_example_image}',
                            'detail': 'low',  # The image is already downsampled to 512x512
                        },
                    },
                ],
            },
            {
                'role': 'assistant',
                'content': """{
  "type": "sail",
  "size": "6.5",
  "brand": "North Spectro",
  "mast_length": "4.92",
  "boom_size": "1.95",
  "year": "N/A",
  "state": "repaired"
}""",
            },
            {
                'role': 'user',
                'content': [
                    {
                        'type': 'text',
                        'text': f"""Convert the following offer into the appropriate JSON format:

Title: {offer.title}
Description: {offer.description}""",
                    },
                    *[
                        {
                            'type': 'image_url',
                            'image_url': {'url': url, 'detail': 'high'},
                        }
                        for url in offer.image_urls
                    ],
                ],
            },
        ],
        temperature=0.0,
        response_format={'type': 'json_object'},
    )

    response_json = response.choices[0].message.content
    if not response_json:
        print('Failed to extract the details of the offer:', offer.title)
        return offer

    try:
        json_data = json.loads(response_json)
        if json_data['type'] == 'N/A':
            return offer

        return DatabaseFactory.parse_parial_entry(json_data, offer)
    except ValueError:
        print('Failed to parse the JSON response:', response_json)
        return offer


async def test():
    offer = await scrape_offer_url(
        'https://www.kleinanzeigen.de/s-anzeige/neil-pryde-segel-windsurfen-7-5-qm/2799076828-230-18675?simcid=ba9ab67c-f24e-4917-80d4-864b57cbba13'
    )

    details = await extract_offer_details(offer)

    pprint(details)
    dump_json(details, 'example_offer_details.json')
    exit()


def list_entries_of_type(entries: list[Entry], type: str) -> list[Entry]:
    return [entry for entry in entries if entry.metadata.type == type]


def to_excel(entries: list[Entry], path: str = 'export.xlsx') -> str:
    wb = Workbook()
    # Remove the default sheet
    if wb.active:
        wb.remove(wb.active)

    for type in 'sail', 'board', 'mast', 'boom', 'full_set', 'full_rig', 'accessory':
        entries_of_type = list_entries_of_type(entries, type)
        if entries_of_type:
            ws: Worksheet = wb.create_sheet(type.capitalize(), 0)
            add_entries_to_worksheet(ws, entries_of_type)

    # Save the workbook
    wb.save(path)

    return path


def add_entries_to_worksheet(ws: Worksheet, entries: list[Entry]) -> None:
    assert len(entries) > 0, 'We need at least one entry to create an Excel sheet'

    headers = list(entries[0].to_excel().keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    max_lengths = [0] * len(headers)

    for row_idx, sail in enumerate(entries, 2):
        for col_idx, value in enumerate(sail.to_excel().values(), 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value.value
            max_lengths[col_idx - 1] = max(max_lengths[col_idx - 1], len(str(value.value)))
            if value.number_format:
                cell.number_format = value.number_format
            if isinstance(value.value, str) and value.value.startswith('http'):
                cell.value = f'=HYPERLINK("{value.value}", "Link")'
                cell.font = Font(color='0000FF', underline='single')

    def rescale_column_width(name: str, width: float) -> None:
        column_index = headers.index(name) + 1
        column_letter = ws.cell(row=1, column=column_index).column_letter
        ws.column_dimensions[column_letter].width = width

    for name, width in zip(headers, max_lengths):
        rescale_column_width(name, width + 4)

    rescale_column_width('Date', 10)
    rescale_column_width('Link', 10)
    rescale_column_width('All other offers', 20)

    # Apply AutoFilter to all columns
    ws.auto_filter.ref = ws.dimensions  # This sets the AutoFilter to cover all the data including headers


async def main():
    if True:
        entries = load_database()

        path = to_excel(entries)
        print(f'Data saved to: {path}')

        exit()

    WINDSURF_SEARCH_URL = BASE_URL + '/s-karlsruhe/seite:{}/windsurfen/k0l9186r50'

    all_offer_links = scrape_all_offer_links_from_search_url(WINDSURF_SEARCH_URL)
    with timeblock('scraping all offers'):
        all_offers = await scrape_all_offers(all_offer_links)

    dump_json(all_offers, CURRENT_OFFERS_FILE)

    with timeblock('loading the database'):
        database_entries = load_database()

    new_offers, old_offers, sold_offers = partition_offers(all_offers, database_entries)

    for entry in sold_offers:
        entry.metadata.offer.sold = True

    with timeblock('updating old offers'):
        for offer, entry in old_offers:
            if offer.title != entry.metadata.offer.title or offer.description != entry.metadata.offer.description:
                # reextract the offer details via llm
                new_entry_details = await extract_offer_details(offer)
                for key, value in new_entry_details.__dict__.items():
                    setattr(entry, key, value)
            # update the entry in the database
            entry.metadata.offer = offer

    # extract the details of the new offers
    with timeblock('extracting the details of the new offers'):
        extracted_details = await asyncio.gather(*[extract_offer_details(offer) for offer in new_offers])

    new_entries = [entry for entry in extracted_details if isinstance(entry, Entry)]
    filtered_out_offers = [offer for offer in extracted_details if isinstance(offer, Offer)]

    dump_json(filtered_out_offers, FILTERED_OUT_OFFERS_FILE)

    # store everything in the database
    new_database_entries = new_entries + database_entries
    dump_json(new_database_entries, DB_FILE)

    # TODO proper display of the available offers (in a web interface?)


if __name__ == '__main__':
    asyncio.run(main())
