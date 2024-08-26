from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font


from src.types import Entry


def list_entries_of_type(entries: list[Entry], type: str) -> list[Entry]:
    return [entry for entry in entries if entry.metadata.type == type]


def to_excel(entries: list[Entry], path: str = 'export.xlsx') -> str:
    wb = Workbook()
    # Remove the default sheet
    if wb.active:
        wb.remove(wb.active)

    for type in 'uninteresting', 'accessory', 'full_rig', 'full_set', 'boom', 'mast', 'board', 'sail':
        entries_of_type = list_entries_of_type(entries, type)
        scraped_on_dict = {entry.metadata.offer.link: entry.to_excel()['Scraped on'].value for entry in entries_of_type}
        entries_of_type.sort(key=lambda entry: scraped_on_dict[entry.metadata.offer.link], reverse=True)
        if entries_of_type:
            ws: Worksheet = wb.create_sheet(type.capitalize(), 0)
            add_entries_to_worksheet(ws, entries_of_type)

    # Save the workbook
    wb.save(path)

    return path


def add_entries_to_worksheet(ws: Worksheet, entries: list[Entry]) -> None:
    assert len(entries) > 0, 'We need at least one entry to create an Excel sheet'

    headers = list(entries[0].to_excel().keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    max_lengths = [len(header) for header in headers]

    for row_idx, entry in enumerate(entries, 2):
        for col_idx, value in enumerate(entry.to_excel().values(), 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value.value
            if isinstance(value.value, float):
                max_lengths[col_idx - 1] = max(max_lengths[col_idx - 1], len(str(round(value.value, 2))))
            else:
                max_lengths[col_idx - 1] = max(max_lengths[col_idx - 1], len(str(value.value)))
            if value.number_format:
                cell.number_format = value.number_format
            if isinstance(value.value, str) and (value.value.startswith('http') or value.value.startswith('C:\\')):
                cell.value = f'=HYPERLINK("{value.value}", "Link")'
                cell.font = Font(color='0000FF', underline='single')

    def rescale_column_width(name: str, width: float) -> None:
        column_index = headers.index(name) + 1
        column_letter = ws.cell(row=1, column=column_index).column_letter
        ws.column_dimensions[column_letter].width = width

    for name, width in zip(headers, max_lengths):
        rescale_column_width(name, width + 4)

    rescale_column_width('Date', 10)
    rescale_column_width('Link', 10)
    rescale_column_width('Images', 10)
    rescale_column_width('All other offers', 20)

    # Apply AutoFilter to all columns
    ws.auto_filter.ref = ws.dimensions  # This sets the AutoFilter to cover all the data including headers
